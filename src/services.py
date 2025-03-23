import json
import logging
from datetime import datetime
import os

import openpyxl

path_to_logs = os.getenv("PATH_TO_LOGS")  # "../logs/app.log"

# Настройка логирования
services_logger = logging.getLogger("services")
# Проверка и создание директории для логов, если она не существует
log_directory = os.path.dirname(path_to_logs)
if not os.path.exists(log_directory):
    os.makedirs(log_directory)
# Обработчик для записи логов в файл
file_handler = logging.FileHandler(path_to_logs, mode="a", encoding="utf-8")
file_formater = logging.Formatter("%(asctime)s - %(name)s: %(funcName)s - %(levelname)s: %(message)s")
file_handler.setFormatter(file_formater)  # Исправлено: передаем file_formater
services_logger.addHandler(file_handler)
# Обработчик для вывода логов в консоль (опционально)
console_handler = logging.StreamHandler()
console_handler.setFormatter(file_formater)
services_logger.addHandler(console_handler)
# Установка уровня логирования
services_logger.setLevel(logging.DEBUG)


def favorable_categories_of_increased_cashback(month, year, path_to_xlsx):
    """Функция для анализа выгодности категорий повышенного кешбэка."""
    try:
        # Чтение Excel-файла
        book = openpyxl.load_workbook(path_to_xlsx)
        sheet = book.active

        # Словарь для хранения результатов
        result_dict = {}

        # Проход по строкам
        for row in range(2, sheet.max_row + 1):
            # Чтение данных из ячеек
            cat_find = sheet.cell(row=row,
                                  column=10).value  # Категория (столбец J)
            pay_find = sheet.cell(
                row=row, column=13
            ).value  # Процент кешбэка (столбец M)
            amount_find = sheet.cell(
                row=row, column=7
            ).value  # Сумма операции (столбец G)
            date_find = sheet.cell(row=row, column=2).value  # Дата (столбец B)

            # Пропустить строку, если данные отсутствуют
            if None in (cat_find, pay_find, amount_find, date_find):
                logging.warning(f"Пропущена строка {row}: отсутствуют данные")
                continue

            # Преобразование даты
            try:
                dt_object = datetime.strptime(date_find, "%d.%m.%Y")
            except ValueError:
                logging.error(f"Ошибка в строке {row}: "
                              f"некорректная дата '{date_find}'")
                continue

            # Проверка месяца и года
            if month == dt_object.month and year == dt_object.year:
                # Вычисление кешбэка
                cashback = round(pay_find * amount_find) * -1

                # Объединяем результаты в один словарь
                if cat_find in result_dict:
                    result_dict[
                        cat_find
                    ] += cashback  # Суммируем кешбэк
                    # для существующей категории

                else:
                    # Добавляем новую категорию
                    result_dict[cat_find] = cashback

        # Преобразуем результат в JSON
        result_json = json.dumps(result_dict, ensure_ascii=False, indent=4)
        return result_json

    except Exception as e:
        logging.error(f"Ошибка при обработке файла: {e}")
        return json.dumps({})  # Возвращаем пустой JSON-объект в случае ошибки
