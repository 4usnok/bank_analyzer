import json
import logging
from datetime import datetime, time

import openpyxl
import pandas as pd
import yfinance as yf

path_to_xlsx = "../data/operations.xlsx"
path_to_cur = "../work_file_cur.json"
# Настройка логирования
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


def greetings() -> None:
    """Функция выводит приветствие, в зависимости от периода времени"""
    dt_now = datetime.now()
    time_now = dt_now.time()  # Получаем текущее время
    # Определяем временные диапазоны
    morning_start = time(6, 0, 0)
    day_start = time(12, 0, 0)
    evening_start = time(18, 0, 0)
    night_start = time(0, 0, 0)
    # Определяем, какое сейчас время суток и выводим соответствующее сообщение
    if night_start <= time_now < morning_start:
        period_of_day = "Доброй ночи"
        return period_of_day
    elif morning_start <= time_now < day_start:
        period_of_day = "Доброе утро"
        return period_of_day
    elif day_start <= time_now < evening_start:
        period_of_day = "Добрый день"
        return period_of_day
    elif evening_start <= time_now or time_now < night_start:
        period_of_day = "Добрый вечер"
        return period_of_day


def for_each_card() -> None:
    """Функция выводит:
    1) последние 4 цифры карты;
    2) общая сумма расходов;
    3) кешбэк (1 рубль на каждые 100 рублей).
    """
    try:
        # Чтение Excel-файла с помощью pandas
        df = pd.read_excel(path_to_xlsx)

        # Логирование доступных столбцов (для отладки)
        logging.info(f"Доступные столбцы в файле: {df.columns.tolist()}")

        # Проверка наличия нужных столбцов
        required_columns = ["Номер карты", "Сумма операции с округлением"]
        mis_col = [col for col in required_columns if col not in df.columns]

        if mis_col:
            raise ValueError(f"Отсутствуют столбцы: {mis_col}")

        # Выбор нужных строк и столбцов
        # Предполагаем, что данные начинаются со второй строки (индекс 1)
        selected_data = df.loc[1:3, required_columns]

        # Преобразуем данные в список словарей
        for_each_card_list = selected_data.apply(
            lambda row: {
                "last_digits": str(
                    row["Номер карты"]
                ),  # Последние 4 цифры номера карты
                "total_spent": row["Сумма операции с округлением"],
                "cashback": round(row["Сумма операции с округлением"] // 100),
            },
            axis=1,
        ).tolist()

        return for_each_card_list
    except Exception as e:
        logging.error(f"Ошибка при чтении файла Excel: {e}")
        return []


def top_trans() -> None:
    """Функция выводит топ-5 транзакций по сумме платежа."""
    # Обработка xlsx файла из которого будем брать необходимые данные
    workbook = openpyxl.load_workbook(path_to_xlsx)
    sheet = workbook.active
    top_trans_list = []
    # Собираем все транзакции в список
    try:
        for row in range(2, sheet.max_row + 1):  # данные со второй строки
            tr_date = sheet.cell(
                row=row, column=2
            ).value  # [2] - номер столбца - "date"
            tr_amount = sheet.cell(
                row=row, column=5
            ).value  # [5] - номер столбца - "amount"
            tr_category = sheet.cell(
                row=row, column=10
            ).value  # [10] - номер столбца - "category"
            tr_description = sheet.cell(
                row=row, column=12
            ).value  # [12] - номер столбца - "description"
            top_trans_list.append(
                {
                    "date": tr_date,
                    "amount": tr_amount,
                    "category": tr_category,
                    "description": tr_description,
                }
            )
        # Сортируем список по убыванию суммы (amount)
        top_trans_list.sort(key=lambda x: x["amount"], reverse=True)
        # Возвращаем топ-5 транзакций
        return top_trans_list[:5]
    except Exception as e:
        logging.error(f"Ошибка при чтении файла Excel: {e}")
        return []


def exchange_rates() -> None:
    """Функция выводит курс валют"""
    exchange_rates_list = []
    # Обработка json файла из которого будем брать
    # необходимые данные апи, чтобы лишний раз не тратить вызовы
    try:
        with open("../work_file_cur.json", "r", encoding="utf-8") as file:
            data_file = json.load(file)
            # Раскроем список
            for row in data_file:
                list_from_base = row.get("base", [])
                list_from_rates = row.get("rates", [])
                # Заполним словарь в списке и добавим его в итоговый
                exchange_rates_list.append(
                    {"currency": list_from_base, "rate": list_from_rates}
                )
        return exchange_rates_list
    except Exception as e:
        logging.error(f"Ошибка при чтении файла JSON: {e}")
        return []


def stock_prices() -> None:
    """Функция выводит стоимость акций из S&P500."""
    try:
        stock_prices_list = []
        # Открываем файл с настройками
        with open("../user_settings.json", "r", encoding="utf-8") as file:
            data_file = json.load(file)
        # Получаем список акций
        list_from_the_dict = data_file.get("user_stocks", [])
        # Проверяем, что список акций не пустой
        if not list_from_the_dict:
            raise ValueError("'user_stocks' пуст или отсутствует в файле.")
        # Формируем список с данными о ценах акций
        for one_stock in list_from_the_dict:
            # Получаем текущую цену акции
            stock_info = yf.Ticker(one_stock)
            current_price = stock_info.history(period="1d")["Close"].iloc[-1]
            # Преобразуем цену в целое число
            current_price_int = int(
                round(current_price)
            )  # Округляем и преобразуем в int
            stock_prices_list.append(
                {
                    "stock": one_stock,
                    "price": current_price_int,  # Цена в виде целого числа
                }
            )
        return stock_prices_list
    except Exception as e:
        logging.error(f"Ошибка при получении цен акций: {e}")
        return []


def main_home_page() -> None:
    """Управляющая функция."""
    json_response = {
        "greeting": greetings(),
        "cards": for_each_card(),
        "top_transactions": top_trans(),
        "currency_rates": exchange_rates(),
        "stock_prices": stock_prices(),
    }
    logging.info("Данные успешно собраны.")
    return json_response


# Пример использования
if __name__ == "__main__":
    try:
        response = main_home_page()
        print(response)
    except Exception as e:
        logging.error(f"Ошибка в главной функции: {e}")
